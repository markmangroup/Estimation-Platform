#!/usr/bin/env python3
"""
Convert Items42.xls (SpreadsheetML XML format) to proper .xlsx format
"""
import openpyxl
import xml.etree.ElementTree as ET
from pathlib import Path
from decimal import Decimal

# Parse the SpreadsheetML XML
xml_file = Path("src/laurel/example_bids/Items42.xls")
output_file = Path("src/laurel/example_bids/Items42.xlsx")

print(f"Reading {xml_file}...")
tree = ET.parse(xml_file)
root = tree.getroot()

# Define namespaces
ns = {
    'ss': 'urn:schemas-microsoft-com:office:spreadsheet',
    'o': 'urn:schemas-microsoft-com:office:office',
    'x': 'urn:schemas-microsoft-com:office:excel',
}

# Create new Excel workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

# Find all worksheets
worksheets = root.findall('.//ss:Worksheet', ns)
print(f"Found {len(worksheets)} worksheet(s)")

for ws_elem in worksheets:
    ws_name = ws_elem.get(f'{{{ns["ss"]}}}Name', 'Sheet1')
    ws = wb.create_sheet(ws_name)
    print(f"  Processing: {ws_name}")

    # Find table
    table = ws_elem.find('.//ss:Table', ns)
    if table is None:
        continue

    # Process rows
    rows = table.findall('.//ss:Row', ns)
    for row_idx, row in enumerate(rows, 1):
        cells = row.findall('.//ss:Cell', ns)
        col_idx = 1

        for cell in cells:
            # Handle Index attribute (skipped columns)
            index_attr = cell.get(f'{{{ns["ss"]}}}Index')
            if index_attr:
                col_idx = int(index_attr)

            # Get cell data
            data_elem = cell.find('.//ss:Data', ns)
            if data_elem is not None:
                cell_value = data_elem.text
                data_type = data_elem.get(f'{{{ns["ss"]}}}Type')

                # Convert types
                if data_type == 'Number' and cell_value:
                    try:
                        cell_value = float(cell_value)
                    except:
                        pass

                ws.cell(row=row_idx, column=col_idx, value=cell_value)

            col_idx += 1

# Save
wb.save(output_file)
print(f"✓ Saved to {output_file}")

# Verify
import pandas as pd
df = pd.read_excel(output_file)
print(f"\n✓ Verification:")
print(f"  Rows: {len(df):,}")
print(f"  Columns: {len(df.columns)}")
print(f"\n  Column names:")
for i, col in enumerate(df.columns, 1):
    print(f"    {i}. {col}")
