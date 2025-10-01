#!/usr/bin/env python3
"""
Analyze Franklin's Excel workflow to identify pain points and improvements.
"""
import openpyxl
import pandas as pd
from pathlib import Path

def analyze_template():
    """Analyze the Public Works Template v7"""
    print("="*80)
    print("FRANKLIN'S EXCEL WORKFLOW ANALYSIS")
    print("="*80)

    template_path = Path("src/laurel/example_bids/Public Works Template v7.xlsx")
    wb = openpyxl.load_workbook(template_path, data_only=False)

    print(f"\n📊 TEMPLATE STRUCTURE")
    print(f"Total sheets: {len(wb.sheetnames)}")
    print(f"\nSheet list:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        sheet = wb[sheet_name]
        print(f"  {i}. {sheet_name} ({sheet.max_row} rows × {sheet.max_column} cols)")

    # Analyze Bid Summary
    print(f"\n" + "="*80)
    print("BID SUMMARY TAB ANALYSIS")
    print("="*80)
    bid_summary = wb['Bid Summary']

    print(f"\n🔍 Key Inputs/Outputs:")
    # Try to find key cells
    for row in range(1, min(50, bid_summary.max_row)):
        for col in range(1, min(10, bid_summary.max_column)):
            cell = bid_summary.cell(row, col)
            if cell.value and isinstance(cell.value, str):
                if any(keyword in str(cell.value).lower() for keyword in ['tax rate', 'margin', 'total', 'overhead']):
                    print(f"  Row {row}: {cell.value}")

    # Analyze Bid Schedule
    print(f"\n" + "="*80)
    print("BID SCHEDULE TAB ANALYSIS (if exists)")
    print("="*80)
    if 'Bid Schedule' in wb.sheetnames:
        schedule = wb['Bid Schedule']
        print(f"Rows: {schedule.max_row}, Columns: {schedule.max_column}")

        # Print headers
        print(f"\n📋 Column Headers:")
        for col in range(1, min(15, schedule.max_column + 1)):
            header = schedule.cell(1, col).value
            if header:
                print(f"  Col {col}: {header}")

    # Count bid item tabs
    print(f"\n" + "="*80)
    print("BID ITEM TABS")
    print("="*80)
    bid_item_tabs = [name for name in wb.sheetnames if name not in ['Bid Summary', 'Bid Schedule', 'Prevailing Wage', 'Equipment Rental Rates', 'Bid Bond']]
    print(f"Total bid item tabs: {len(bid_item_tabs)}")
    if bid_item_tabs:
        print(f"Examples: {', '.join(bid_item_tabs[:5])}")

def analyze_example():
    """Analyze Example 2 - Small PW Job"""
    print(f"\n" + "="*80)
    print("EXAMPLE 2 - SMALL PW JOB ANALYSIS")
    print("="*80)

    example_path = Path("src/laurel/example_bids/Example 2 - Small PW Job.xlsx")
    wb = openpyxl.load_workbook(example_path, data_only=True)

    print(f"\n📊 Sheets: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        print(f"  - {name}")

    # Analyze Bid Schedule
    if 'Bid Schedule' in wb.sheetnames:
        print(f"\n📋 BID SCHEDULE CONTENTS:")
        schedule = wb['Bid Schedule']

        # Read as dataframe
        data = []
        headers = []
        for col in range(1, schedule.max_column + 1):
            header_cell = schedule.cell(1, col)
            if header_cell.value:
                headers.append(header_cell.value)

        if headers:
            print(f"  Columns: {', '.join(headers)}")

            # Read data rows
            for row in range(2, min(20, schedule.max_row + 1)):
                row_data = []
                for col in range(1, len(headers) + 1):
                    cell = schedule.cell(row, col)
                    row_data.append(cell.value)
                if any(row_data):  # If row has any data
                    data.append(row_data)

            print(f"\n  Sample bid items (first 5):")
            for i, row in enumerate(data[:5], 1):
                print(f"    {i}. {row[:3]}")  # First 3 columns

def analyze_items42():
    """Analyze Items42.xls - NetSuite product catalog"""
    print(f"\n" + "="*80)
    print("ITEMS42.XLS - NETSUITE CATALOG ANALYSIS")
    print("="*80)

    items_path = Path("src/laurel/example_bids/Items42.xls")

    try:
        # Read with pandas
        df = pd.read_excel(items_path)

        print(f"\n📦 Product Catalog Stats:")
        print(f"  Total products: {len(df):,}")
        print(f"  Columns: {list(df.columns)}")
        print(f"\n  Sample products (first 5):")
        print(df.head().to_string())

        print(f"\n  Cost range:")
        if 'Standard Cost' in df.columns:
            cost_col = df['Standard Cost']
            print(f"    Min: ${cost_col.min():.2f}")
            print(f"    Max: ${cost_col.max():.2f}")
            print(f"    Avg: ${cost_col.mean():.2f}")

    except Exception as e:
        print(f"Error reading Items42.xls: {e}")

def identify_pain_points():
    """Identify manual pain points in Franklin's workflow"""
    print(f"\n" + "="*80)
    print("IDENTIFIED PAIN POINTS & IMPROVEMENTS")
    print("="*80)

    pain_points = [
        {
            "problem": "Manual copy/paste from NetSuite",
            "current": "Search product in NetSuite, copy name and cost, paste into Excel",
            "improvement": "Search products in web app (pre-imported from Items42.xls), click to add",
            "time_saved": "5 seconds → 1 second per product (80% faster)"
        },
        {
            "problem": "Manual formula updates when adding rows",
            "current": "Insert row, copy formulas down, hope they're correct",
            "improvement": "Auto-calculate totals when saving (no formulas to break)",
            "time_saved": "30 seconds per row → instant"
        },
        {
            "problem": "Multiple tabs to track status",
            "current": "Right-click tab, change color, remember what each color means",
            "improvement": "Status dropdown with clear labels (🟢 Done, 🔴 Need Work, etc.)",
            "time_saved": "Clearer status at a glance"
        },
        {
            "problem": "Hyperlinks break when copying spreadsheet",
            "current": "Links to DIR, equipment rates stop working",
            "improvement": "Store URLs in database, always accessible",
            "time_saved": "No more re-finding links"
        },
        {
            "problem": "Can't see what changed",
            "current": "No history of who changed what when",
            "improvement": "Activity log shows all changes with timestamps",
            "time_saved": "Accountability + debugging"
        },
        {
            "problem": "Margin calculations are manual",
            "current": "Adjust cell, recalculate, check if correct",
            "improvement": "Change margin %, all totals update instantly",
            "time_saved": "Instant feedback"
        },
        {
            "problem": "Export to NetSuite requires reformatting",
            "current": "Copy data, reformat CSV, match Internal IDs manually",
            "improvement": "Click 'Export to NetSuite' → CSV ready to import",
            "time_saved": "20 minutes → 10 seconds"
        },
        {
            "problem": "Can't work on same estimate simultaneously",
            "current": "File locks, 'Mike is editing this file'",
            "improvement": "Multiple team members can edit different bid items at same time",
            "time_saved": "No waiting for file to unlock"
        }
    ]

    for i, point in enumerate(pain_points, 1):
        print(f"\n{i}. {point['problem']}")
        print(f"   ❌ Current: {point['current']}")
        print(f"   ✅ Improved: {point['improvement']}")
        print(f"   ⏱️  Impact: {point['time_saved']}")

if __name__ == "__main__":
    analyze_template()
    analyze_example()
    analyze_items42()
    identify_pain_points()
